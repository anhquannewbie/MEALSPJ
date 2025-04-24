#!/usr/bin/env python3
# process.py

import sys, glob, os
import pandas as pd
from openpyxl import load_workbook

def next_name(base):
    """
    Tìm các file base{n}.xlsx, trả về base{m}.xlsx với m = max(n)+1 hoặc 1 nếu chưa có.
    """
    pattern = f"{base}[0-9]*.xlsx"
    files = glob.glob(pattern)
    nums = []
    for fn in files:
        stem = os.path.splitext(os.path.basename(fn))[0]
        num = stem[len(base):]
        if num.isdigit():
            nums.append(int(num))
    n = max(nums) + 1 if nums else 1
    return f"{base}{n}.xlsx"

def split_src(src):
    wb0 = load_workbook(src)
    ws0 = wb0.active

    # 1) tìm row “SÁNG” và “TRƯA”
    br = lu = None
    for r in range(1, ws0.max_row+1):
        v = ws0.cell(r,1).value
        if isinstance(v,str):
            U = v.upper()
            if 'BẢNG CHẤM ĂN SÁNG' in U: br = r
            if 'BẢNG CHẤM CƠM TRƯA' in U: lu = r

    def find_hdr(st):
        for rr in range(st+1, st+6):
            if ws0.cell(rr,1).value == 'STT':
                return rr
        return None

    br_h = find_hdr(br)
    lu_h = find_hdr(lu) if lu else None

    # 2) lớp từ tiêu đề SÁNG
    raw = ws0.cell(br,1).value
    classes = [c.strip() for c in raw.split('LỚP:')[-1].split(',') if c.strip()]

    # 3) tìm cột ngày (header là số)
    def days(hr):
        arr=[]
        for c in range(1, ws0.max_column+1):
            val = ws0.cell(hr,c).value
            if isinstance(val,(int,float)) or (isinstance(val,str) and val.strip().isdigit()):
                arr.append(c)
        return arr

    br_days = days(br_h)
    lu_days = days(lu_h) if lu_h else []

    regs = [
        ('s', br, br_h, br_days, (lu-1) if lu else ws0.max_row),
    ]
    if lu:
        regs.append(('t', lu, lu_h, lu_days, ws0.max_row))

    out_files = []
    for prefix, trow, hrow, dcols, endr in regs:
        for cls in classes:
            wb = load_workbook(src)
            ws = wb.active

            # xóa row trước vùng
            for r in range(trow-1, 0, -1):
                ws.delete_rows(r)
            # xóa row sau vùng
            shift = trow-1
            end_s = endr - shift
            for r in range(ws.max_row, end_s, -1):
                ws.delete_rows(r)
            # xóa row không cùng lớp (col D=4)
            new_hdr = hrow - shift
            for r in range(ws.max_row, new_hdr, -1):
                v = ws.cell(r,4).value
                if v is None or str(v).strip().lower()!=cls.lower():
                    ws.delete_rows(r)

            base = f"{prefix}{cls.replace(' ','').lower()}"
            out = next_name(base)
            wb.save(out)
            out_files.append(out)
            print("Split→", out)
    return out_files

def clean_files(files):
    for f in files:
        df = pd.read_excel(f, skiprows=2, engine='openpyxl')
        # xóa cột D
        df.drop(df.columns[3], axis=1, inplace=True)
        # replace 0→KP trong col ngày
        day_cols = [c for c in df.columns if str(c).strip().isdigit()]
        for c in day_cols:
            df[c] = df[c].apply(lambda x: 'KP' if x==0 else x)
        df.to_excel(f, index=False)
        print("Clean→", f)

def main():
    if len(sys.argv)==2:
        src = sys.argv[1]
        # nếu là nguồn gốc (có “BẢNG…” ở A1) thì split, rồi clean
        wb = load_workbook(src, read_only=True)
        A1 = wb.active.cell(1,1).value or ''
        if isinstance(A1,str) and 'BẢNG' in A1.upper():
            out = split_src(src)
            clean_files(out)
            return
        # nếu là file đầu ra thì chỉ clean
        clean_files([src])
    else:
        # không args → clean các file s*.xlsx t*.xlsx
        files = glob.glob('s*.xlsx') + glob.glob('t*.xlsx')
        clean_files(files)

if __name__=="__main__":
    main()
