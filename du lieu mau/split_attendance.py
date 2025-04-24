#!/usr/bin/env python3
# split_attendance.py

import sys
from openpyxl import load_workbook

def split_with_format(src):
    wb0 = load_workbook(src)
    ws0 = wb0.active

    # 1) Tìm row tiêu đề SÁNG và TRƯA (dựa vào cột A chứa text)
    br_title = lu_title = None
    for r in range(1, ws0.max_row+1):
        cell = ws0.cell(r,1).value
        if isinstance(cell, str):
            u = cell.upper()
            if 'BẢNG CHẤM ĂN SÁNG' in u:
                br_title = r
            if 'BẢNG CHẤM CƠM TRƯA' in u:
                lu_title = r
    if not br_title or not lu_title:
        print("Không tìm thấy 2 tiêu đề SÁNG & TRƯA.")
        sys.exit(1)

    # 2) Tìm header row (STT...) ngay sau tiêu đề
    def find_header(start_row):
        for r in range(start_row+1, start_row+6):
            if ws0.cell(r,1).value == 'STT':
                return r
        return None

    br_header = find_header(br_title)
    lu_header = find_header(lu_title)
    if not br_header or not lu_header:
        print("Không tìm thấy dòng header STT sau tiêu đề.")
        sys.exit(1)

    # 3) Cột “Lớp” luôn ở D (col=4)
    class_col = 4

    # 4) Xác định ngày trong mỗi header (số 1..31)
    def get_days(header_row):
        days = []
        for c in range(1, ws0.max_column+1):
            v = ws0.cell(header_row, c).value
            if isinstance(v,(int,float)) or (isinstance(v,str) and v.strip().isdigit()):
                days.append(c)
        return days

    br_days = get_days(br_header)
    lu_days = get_days(lu_header)

    # 5) Lấy danh sách lớp từ chính dòng tiêu đề SÁNG sau "LỚP:"
    raw = ws0.cell(br_title,1).value
    classes = [x.strip() for x in raw.split('LỚP:')[-1].split(',') if x.strip()]

    # 6) Với mỗi ca (s/t) và mỗi lớp, copy sheet & xóa row ngoài vùng & khác lớp
    for prefix, (title_row, header_row, days, end_after) in {
        's': (br_title, br_header, br_days, lu_title-1),
        't': (lu_title, lu_header, lu_days, ws0.max_row)
    }.items():
        for cls in classes:
            wb = load_workbook(src)
            ws = wb.active

            # -- Xóa tất cả row trước vùng start
            for r in range(title_row-1, 0, -1):
                ws.delete_rows(r)
            # -- Xóa tất cả row sau vùng end (lưu ý shift index)
            shift = title_row-1
            end_shifted = end_after - shift
            for r in range(ws.max_row, end_shifted, -1):
                ws.delete_rows(r)

            # -- Xóa row không cùng lớp
            # header' = header_row - shift
            header_new = header_row - shift
            for r in range(ws.max_row, header_new, -1):
                val = ws.cell(r, class_col).value
                if val is None or str(val).strip().lower() != cls.lower():
                    ws.delete_rows(r)

            # -- Lưu file
            safe = cls.replace(' ','').lower()
            out = f"{prefix}{safe}.xlsx"
            wb.save(out)
            print("Saved:", out)

if __name__=="__main__":
    if len(sys.argv)!=2:
        print("Usage: python split_attendance.py filegoc.xlsx")
        sys.exit(1)
    split_with_format(sys.argv[1])
