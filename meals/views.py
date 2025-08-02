from django.shortcuts import render, get_object_or_404, redirect
from .models import MealRecord
from .forms import MealRecordForm
from django.contrib import messages
from django.utils.translation import gettext as _
from django.http import JsonResponse
from .models import Student
from django.views.decorators.csrf import csrf_exempt
from datetime import date, timedelta,datetime
from .models import MealRecord, Student
from .models import StudentPayment,MealPrice,AuditLog
from .forms import StudentPaymentForm,ClassRoom
from django.db import connection
import json
from django.db.models import Q
import calendar
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font
from decimal import Decimal
from openpyxl.utils import get_column_letter
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.contrib.auth import authenticate, login as auth_login
from django.urls import reverse
from django.contrib.auth.decorators import login_required, permission_required
from django.core.exceptions import PermissionDenied
from .utils import get_current_price
from django.db import transaction
from django.utils.dateparse import parse_date
from django.db.models.functions import ExtractMonth, ExtractYear

def admin_logout_view(request):
    """Handle both GET and POST logout requests for admin"""
    from django.contrib.auth import logout
    logout(request)
    return redirect('/login/')

@csrf_exempt
def logout_view(request):
    """Handle both GET and POST logout requests"""
    logout(request)
    return redirect('login')
def user_login(request):
    error = None
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            auth_login(request, user)
            # nếu là staff (admin site) thì redirect luôn vào /admin/
            if user.is_staff:
                return redirect(reverse('admin:index'))
            # ngược lại về home
            return redirect('home')
        else:
            error = "Sai tên đăng nhập hoặc mật khẩu"
    return render(request, 'login.html', {'error': error})
@login_required(login_url='login')
@permission_required('meals.view_statistics', raise_exception=True)
def statistics_print_view(request):
    mode = request.GET.get('mode', 'month')
    selected_term = request.GET.get('term')
    selected_class_id = request.GET.get('class_id')
    selected_month = request.GET.get('month')
    selected_start_month = request.GET.get('start_month')
    selected_end_month = request.GET.get('end_month')

    context = {
        'mode': mode,
        'selected_term': selected_term,
        'selected_class_id': selected_class_id,
        'selected_month': selected_month,
        'selected_start_month': selected_start_month,
        'selected_end_month': selected_end_month,
    }

    if mode == 'month' and selected_term and selected_month and selected_class_id:
        m_str, y_str = selected_month.split('/')
        month_i, year_i = int(m_str), int(y_str)
        max_day = calendar.monthrange(year_i, month_i)[1]
        students_qs = Student.objects.filter(classroom_id=int(selected_class_id))
        students = sorted(
            students_qs,
            key=lambda s: s.name.strip().split()[-1].upper()
        )

        data_breakfast = []
        data_lunch = []
        totals_lunch = [0] * max_day
        totals_meals_sum = 0
        totals_food_cost_sum = 0
        totals_tuition_fee_sum = 0
        totals_due_sum = 0

        for meal_type in ['Bữa sáng', 'Bữa trưa']:
            rows = []
            for stu in students:
                marks = []
                for d in range(1, max_day + 1):
                    rec = MealRecord.objects.filter(
                        student=stu,
                        date__year=year_i,
                        date__month=month_i,
                        date__day=d,
                        meal_type=meal_type
                    ).first()
                    if rec:
                        if rec.status == "Đủ":
                            mark = "X"
                        elif rec.status == "Thiếu" and rec.non_eat == 2:
                            mark = "KP"
                        elif rec.status == "Thiếu" and rec.non_eat == 1:
                            mark = "P"
                        else:
                            mark = "0"
                    else:
                        mark = "0"
                    marks.append(mark)
                    if meal_type == 'Bữa trưa' and mark in ["X", "KP"]:
                        totals_lunch[d-1] += 1
                total = sum(1 for m in marks if m in ["X", "KP"])
                row = {'student': stu, 'marks': marks, 'total': total}
                if meal_type == 'Bữa trưa':
                    ym = f"{year_i}-{month_i:02d}"
                    sp = StudentPayment.objects.filter(student=stu, month=ym).first()
                    if sp:
                        br_count = MealRecord.objects.filter(student=stu, date__year=year_i, date__month=month_i, meal_type='Bữa sáng').count()
                        lu_count = total
                        row['food_cost'] = br_count * sp.meal_price.breakfast_price + lu_count * sp.meal_price.lunch_price
                        row['tuition_fee'] = sp.tuition_fee
                        row['total_due'] = row['food_cost'] + row['tuition_fee']
                        totals_meals_sum += total
                        totals_food_cost_sum += row['food_cost']
                        totals_tuition_fee_sum += row['tuition_fee']
                        totals_due_sum += row['total_due']
                rows.append(row)
            if meal_type == 'Bữa sáng':
                data_breakfast = rows
            else:
                data_lunch = rows

        context.update({
            'header_days': list(range(1, max_day + 1)),
            'data_breakfast': data_breakfast,
            'data_lunch': data_lunch,
            'totals_lunch': totals_lunch,
            'totals_meals_sum': totals_meals_sum,
            'totals_food_cost_sum': totals_food_cost_sum,
            'totals_tuition_fee_sum': totals_tuition_fee_sum,
            'totals_due_sum': totals_due_sum,
        })

    elif mode == 'year' and selected_term and selected_start_month and selected_end_month and selected_class_id:
        start_month, start_year = map(int, selected_start_month.split('/'))
        end_month, end_year = map(int, selected_end_month.split('/'))
        months = []
        y, m = start_year, start_month
        while (y < end_year) or (y == end_year and m <= end_month):
            months.append((y, m))
            m += 1
            if m > 12:
                m = 1
                y += 1

        groups = [months[i:i+5] for i in range(0, len(months), 5)]
        data_year = []
        students_qs = Student.objects.filter(classroom_id=int(selected_class_id))
        students = sorted(
            students_qs,
            key=lambda s: s.name.strip().split()[-1].upper()
        )

        for group in groups:
            group_data = []
            group_totals = [[0, 0, 0] for _ in group]
            for stu in students:
                # ---- mỗi tháng trả về 3 giá trị: paid, spent, remaining ----
                cell_list = []
                for idx, (yy, mm) in enumerate(group):
                    ym = f"{yy}-{mm:02d}"
                    sp = StudentPayment.objects.filter(student=stu, month=ym).first()
                    if sp:
                        paid      = sp.amount_paid
                        tuition   = sp.tuition_fee
                        remaining = sp.remaining_balance or Decimal('0')
                        price     = sp.meal_price
                        fee_b     = Decimal(price.breakfast_price)
                        fee_l     = Decimal(price.lunch_price)
                    else:
                        paid      = tuition = remaining = Decimal('0')
                        fee_b     = fee_l   = Decimal('0')
                    recs    = MealRecord.objects.filter(
                        student=stu, date__year=yy, date__month=mm,
                        meal_type__in=["Bữa sáng","Bữa trưa"]
                    )
                    spent_meals = sum(
                        fee_b if r.meal_type=="Bữa sáng" else fee_l
                        for r in recs
                        if (r.status=='Đủ' or (r.status=='Thiếu' and r.non_eat==2))
                    )
                    spent = tuition + spent_meals
                    group_totals[idx][0] += int(paid)
                    group_totals[idx][1] += int(spent)
                    group_totals[idx][2] += int(remaining)
                    cell_list.append({
                        'paid':      f"{int(paid):,}",
                        'spent':     f"{int(spent):,}",
                        'remaining': f"{int(remaining):,}",
                    })
                group_data.append({'student': stu, 'cells': cell_list})
                group_totals_fmt = [[f"{x:,}" for x in triple] for triple in group_totals]
            data_year.append({
                'title': f"{group[0][1]}/{group[0][0]} - {group[-1][1]}/{group[-1][0]}",
                'data': group_data,
                'months': group,
                'totals': group_totals_fmt, 
            })

        context['data_year'] = data_year

    return render(request, 'meals/statistics_print.html', context)

# Other existing views remain unchanged (omitted for brevity)
@login_required(login_url='login')
@permission_required('meals.change_studentpayment', raise_exception=True)
def student_payment_edit(request, pk=None):
    classrooms = ClassRoom.objects.all()

    # Nếu sửa (pk có), load bản ghi; nếu thêm (pk=None), payment=None
    payment = get_object_or_404(StudentPayment, pk=pk) if pk else None

    if request.method == "POST":
        student_id = request.POST.get('student')
        month_val  = request.POST.get('month')

        # --- 1) Tìm duplicate khác nếu có (cùng student+month nhưng khác pk hiện tại) ---
        dup_qs = StudentPayment.objects.filter(student_id=student_id, month=month_val)
        if payment:
            dup_qs = dup_qs.exclude(pk=payment.pk)
        duplicate = dup_qs.first()

        # --- 2) Chọn instance để save: nếu có duplicate thì ghi đè lên nó ---
        form_instance = duplicate if duplicate else payment
        form = StudentPaymentForm(request.POST, instance=form_instance)

        if form.is_valid():
            saved = form.save()
            AuditLog.objects.create(
                user    = request.user,
                action  = 'change',
                path    = request.path,
                data    = json.dumps({
                    'model':           'studentpayment',
                    'object_id':       saved.pk,
                    'student_name':    saved.student.name,
                    'classroom_name':  saved.student.classroom.name,
                    'tuition_fee': format(saved.tuition_fee, ",.2f"),
                    'amount_paid': format(saved.amount_paid, ",.2f"),
                    'meal_price':  format(saved.meal_price.daily_price, ",.2f") if saved.meal_price else "0.00",
                    'month':           month_val,
                }, ensure_ascii=False)
            )
            # Nếu đang edit một bản mà ghi đè lên duplicate thì xóa bản cũ ban đầu
            if duplicate and payment and duplicate.pk != payment.pk:
                payment.delete()

            # --- 3) (Tuỳ chọn) Re-calc remaining_balance cho tất cả tháng kế tiếp ---
            # 1) Lấy list và tính dần dư nợ
            payments = list(
                StudentPayment.objects
                .filter(student=saved.student)
                .order_by('month')
            )
            prior = Decimal('0')
            for p in payments:
                # tính tổng tiền ăn như trước
                year, mon = map(int, p.month.split('-'))
                recs = MealRecord.objects.filter(
                    student=p.student,
                    date__year=year, date__month=mon,
                    meal_type__in=['Bữa sáng','Bữa trưa']
                ).filter(
                    Q(status='Đủ') | Q(status='Thiếu', non_eat=2)
                )
                meal_cost = sum(
                    p.meal_price.breakfast_price if r.meal_type=='Bữa sáng' else p.meal_price.lunch_price
                    for r in recs
                )
                p.remaining_balance = (
                    p.amount_paid
                    - p.tuition_fee
                    - meal_cost
                    +prior
                )
                prior = p.remaining_balance

            # 2) Ghi một lần vào DB, tránh lock
            with transaction.atomic():
                StudentPayment.objects.bulk_update(payments, ['remaining_balance'])

            messages.success(request, "Lưu thành công")
            return redirect(request.path)
        else:
            # gom lỗi rồi hiển thị
            errs = sum(form.errors.values(), [])
            messages.error(request, "Lỗi: " + "; ".join(errs))

    else:
        form = StudentPaymentForm(instance=payment)

    # Tính số bữa để hiển thị dưới form
    breakfast_count = lunch_count = 0
    if form.instance and form.instance.pk:
        y, mon = map(int, form.instance.month.split('-'))
        breakfast_count = MealRecord.objects.filter(
            student=form.instance.student,
            date__year=y, date__month=mon,
            meal_type='Bữa sáng'
        ).filter(Q(status='Đủ')|Q(status='Thiếu', non_eat=2)).count()
        lunch_count = MealRecord.objects.filter(
            student=form.instance.student,
            date__year=y, date__month=mon,
            meal_type='Bữa trưa'
        ).filter(Q(status='Đủ')|Q(status='Thiếu', non_eat=2)).count()

    return render(request, 'payments/student_payment_form.html', {
        'form':            form,
        'classrooms':      classrooms,
        'is_edit':         bool(payment),
        'breakfast_count': breakfast_count,
        'lunch_count':     lunch_count,
    })

# Các view khác giữ nguyên
def ajax_load_students(request):
    classroom_id = request.GET.get('classroom_id')
    if classroom_id:
        qs = Student.objects.filter(classroom_id=classroom_id)
        students = sorted(
            qs,
            key=lambda s: s.name.strip().split()[-1].upper()
        )
        student_list = [{'id': s.id, 'name': s.name} for s in students]
    else:
        student_list = []
    return JsonResponse(student_list, safe=False)

def ajax_load_previous_balance(request):
    student_id = request.GET.get('student_id')
    month = request.GET.get('month')  # 'YYYY-MM'
    prior_balance = 0
    if student_id and month:
        from datetime import datetime, timedelta
        dt = datetime.strptime(month, '%Y-%m')
        prev_month = (dt.replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
        
        from .models import StudentPayment
        prev_payment = StudentPayment.objects.filter(student_id=student_id, month=prev_month).order_by('-id').first()
        if prev_payment:
            prior_balance = prev_payment.remaining_balance or 0
    return JsonResponse({'prev_month_balance': prior_balance})

def ajax_load_payment_details(request):
    student_id = request.GET.get('student_id')
    month = request.GET.get('month')  # format 'YYYY-MM'
    data = {}

    if student_id and month:
        # 1. Lấy payment của THÁNG HIỆN TẠI (nếu có)
        payment = StudentPayment.objects.filter(student_id=student_id, month=month).first()
        if payment:
            # có rồi thì lấy y nguyên
            price = payment.meal_price
            data['meal_price']           = str(price.pk)
            data['tuition_fee']          = str(payment.tuition_fee)
            data['daily_price']          = str(price.daily_price)
            data['breakfast_price']      = str(price.breakfast_price)
            data['lunch_price']          = str(price.lunch_price)
            data['amount_paid']          = str(payment.amount_paid)
            data['current_month_payment']= str(payment.remaining_balance)
        else:
            # Chưa có record tháng này -> load cấu hình & học phí từ bản ghi gần nhất
            prev_cfg = (
                StudentPayment.objects
                .filter(student_id=student_id, month__lt=month)
                .order_by('-month')
                .first()
            )
            if prev_cfg:
                price = prev_cfg.meal_price
                data['meal_price']      = str(price.pk)
                data['tuition_fee']     = str(prev_cfg.tuition_fee)
                data['daily_price']     = str(price.daily_price)
                data['breakfast_price'] = str(price.breakfast_price)
                data['lunch_price']     = str(price.lunch_price)
            else:
                # Không có lần nào trước đó
                data['meal_price']      = ''
                data['tuition_fee']     = ''
                data['daily_price']     = ''
                data['breakfast_price'] = ''
                data['lunch_price']     = ''
            # Các ô tiền đóng vẫn để trống
            data['amount_paid']         = ''
            data['current_month_payment'] = ''

        # 2. Tính "Tiền tháng trước"
        from datetime import datetime, timedelta
        dt = datetime.strptime(month, '%Y-%m')
        prev_month = (dt.replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
        prev_payment = StudentPayment.objects.filter(student_id=student_id, month=prev_month).first()
        if prev_payment:
            data['prev_month_balance'] = str(prev_payment.remaining_balance)
        else:
            data['prev_month_balance'] = '0'

        # 3. Tính số bữa sáng và bữa trưa (breakfast_count, lunch_count)
        # Lấy year, month từ tham số
        year_val = dt.year
        month_val = dt.month

        breakfast_count = 0
        lunch_count = 0

        # Lấy các MealRecord của học sinh trong year_val, month_val
        from django.db.models import Q
        meal_records = MealRecord.objects.filter(
            student_id=student_id,
            date__year=year_val,
            date__month=month_val
        )
        # Duyệt qua meal_records, đếm bữa sáng và trưa nếu status="Đủ" hoặc (status="Thiếu" và non_eat=2)
        for r in meal_records:
            if (r.status == "Đủ") or (r.status == "Thiếu" and r.non_eat == 2):
                if r.meal_type == "Bữa sáng":
                    breakfast_count += 1
                elif r.meal_type == "Bữa trưa":
                    lunch_count += 1

        data['breakfast_count'] = breakfast_count
        data['lunch_count'] = lunch_count

    else:
        # Nếu thiếu student hoặc month -> mọi thứ để trống
        data['tuition_fee'] = ''
        data['daily_price'] = ''
        data['amount_paid'] = ''
        data['current_month_payment'] = ''
        data['prev_month_balance'] = '0'
        data['breakfast_count'] = 0
        data['lunch_count'] = 0

    return JsonResponse(data)

def bulk_meal_record_save(request):
    
    if request.method == 'POST':
        class_name = request.POST.get('class_name')
        meal_type = request.POST.get('meal_type')
        student_ids = request.POST.getlist('student_ids[]')
        print("DEBUG: student_ids =", student_ids)  # Debug: in ra danh sách học sinh có tick
        date_str = request.POST.get('record_date') or request.POST.get('date')
        record_date = parse_date(date_str) if date_str else date.today()

        absence_data_str = request.POST.get('absence_data', '{}')
        absence_data = json.loads(absence_data_str)
        reason_data_str  = request.POST.get('reason_data', '{}')
        reason_data      = json.loads(reason_data_str)
        print("DEBUG: absence_data =", absence_data)  # Debug: in ra dữ liệu dropdown

        students = Student.objects.filter(classroom__name=class_name)
        MealRecord.objects.filter(student__in=students, date=record_date, meal_type=meal_type).delete()

        for student in students:
            sid = str(student.id)
            if sid in student_ids:
                status = "Đủ"
                non_eat = 0
            else:
                status = "Thiếu"
                non_eat = int(absence_data.get(sid, "2"))
            print("DEBUG: For student", sid, "status =", status, "non_eat =", non_eat)
            rec = MealRecord.objects.create(
                student=student,
                date=record_date,
                meal_type=meal_type,
                status=status,
                non_eat=non_eat,
                absence_reason=reason_data.get(sid, '').strip()
            )
            action = 'change' 
            # Log ngay sau khi tạo record
            obj_disp = (
                f"Bản ghi bữa ăn - {rec.date.strftime('%d/%m/%Y')} - "
                f"Bản ghi bữa ăn - {student.name} - "
                f"{student.classroom.name} - {rec.get_meal_type_display()}"
                
            )
            AuditLog.objects.create(
                user=request.user,
                action=action,           # <- dùng biến action ở đây
                path=request.path,
                data=json.dumps({
                    # model_name phải là 'mealrecord' để khớp đúng nhánh bên Admin
                    'model':          rec._meta.model_name,
                    'object_id':      rec.pk,
                    'object_display': obj_disp
                }, ensure_ascii=False)
            )
            current_month = record_date.strftime("%Y-%m")
            try:
                sp = StudentPayment.objects.get(student=student, month=current_month)
                sp.save()  # chạy lại logic tính non_eat tổng, remaining_balance,...
            except StudentPayment.DoesNotExist:
                # nếu chưa có thì bỏ qua hoặc log warning
                continue
        student_names = [stu.name for stu in students]
        
        return JsonResponse({"message": "success"}, status=200)
    return JsonResponse({"error": "Invalid method"}, status=400)

def bulk_meal_record_create(request):
    # 1) Lấy distinct list các term (Học kỳ/Niên khóa), mới nhất ở đầu
    terms_list = ClassRoom.objects.values_list('term', flat=True).distinct().order_by('-term')
    # 2) Nếu có ?term=… từ URL thì dùng, không thì default là phần tử đầu (latest)
    selected_term = request.GET.get('term') or (terms_list[0] if terms_list else '')
    # 2) Nếu user đã chọn term thì load lớp; còn không thì để rỗng
    if selected_term:
        # Lấy cả id và name để JS/HTML dễ render <option value="{{ id }}">{{ name }}</option>
        class_list = ClassRoom.objects \
            .filter(term=selected_term) \
            .order_by('name') \
            .values('id', 'name')
    else:
        class_list = []

    return render(request, 'meals/bulk_meal_form.html', {
        'terms_list':    terms_list,
        'selected_term': selected_term,
        'class_list':    class_list,
        'title':         'Nhập Dữ liệu Bữa Ăn',
        'default_date':  date.today().isoformat(),
        'is_superadmin': request.user.is_superuser,
    })

def meal_record_list(request):
    records = MealRecord.objects.all().order_by('-date')
    return render(request, 'meals/meal_list.html', {'records': records, 'title': _("Danh sách Bữa Ăn")})

def meal_record_detail(request, pk):
    record = get_object_or_404(MealRecord, pk=pk)
    return render(request, 'meals/meal_detail.html', {'record': record, 'title': _("Chi tiết Bữa Ăn")})

def load_students(request):
    # Ajax view: Lấy danh sách học sinh dựa theo lớp học đã chọn
    class_name = request.GET.get('class_name')
    students = []
    if class_name:
        qs = Student.objects.filter(classroom__name=class_name)
        students = sorted(
            qs,
            key=lambda s: s.name.strip().split()[-1].upper()
        )
    data = [{'id': student.id, 'name': student.name} for student in students]
    return JsonResponse(data, safe=False)
def ajax_load_mealdata(request):
    class_name = request.GET.get('class_name')
    meal_type  = request.GET.get('meal_type')
    date_str = request.GET.get('date')
    # Nếu có date_str thì parse, không thì dùng today
    today = parse_date(date_str) if date_str else date.today()
    qs = Student.objects.filter(classroom__name=class_name)
    students = sorted(
        qs,
        key=lambda s: s.name.strip().split()[-1].upper()
    )

    data = []
    for student in students:
        record = MealRecord.objects.filter(
            student=student, date=today, meal_type=meal_type
        ).first()

        if record:
            item = {
                'id':      student.id,
                'name':    student.name,
                'checked': (record.status == 'Đủ'),
                'non_eat': record.non_eat,
                'reason':  record.absence_reason or ''
            }
        else:
            # mặc định: tick Ăn đủ nhưng cho phép chỉnh non_eat=1 và reason=''
            item = {
                'id':      student.id,
                'name':    student.name,
                'checked': True,
                'non_eat': 1,
                'reason':  ''
            }
        data.append(item)

    return JsonResponse(data, safe=False)
@login_required(login_url='login')
def statistics_view(request):
    from_bulk = (request.GET.get('from') == 'bulk')
    mode = request.GET.get('mode', 'month')

    # 3. Nếu từ bulk thì ẩn year + cột tiền
    hide_year = from_bulk
    hide_financial = from_bulk
    # Danh sách lớp và năm
    # --- MỚI: Lấy năm từ ClassRoom thay vì MealRecord ---
    # Đọc selected_year trước để filter lớp
    selected_term = request.GET.get('term')
    # Lấy danh sách năm duy nhất từ ClassRoom.year, sắp xếp giảm dần
    terms_list    = ClassRoom.objects.values_list('term', flat=True).distinct().order_by('-term')

    # Lấy danh sách lớp (ClassRoom) chỉ thuộc năm đã chọn (nếu selected_year đúng); nếu không, trả về rỗng
    if selected_term:
        try:
            classrooms = ClassRoom.objects.filter(term=selected_term).order_by('name')
        except ValueError:
            classrooms = ClassRoom.objects.none()
    else:
        classrooms = ClassRoom.objects.none()

    # Thời điểm này, đã có years_list và classrooms lọc theo năm

    #Đọc chung params
    mode               = request.GET.get('mode', 'month')
    #selected_year      = request.GET.get('year')
    selected_month     = request.GET.get('month')
    selected_meal_type = request.GET.get('meal_type')
    selected_class_id  = request.GET.get('class_id')

    # Context cơ bản
    context = {
        'classrooms':         classrooms,
        'terms_list':        terms_list,
        'mode':               mode,
        'selected_term':     selected_term,
        'selected_month':     selected_month,
        'selected_meal_type': selected_meal_type,
        'selected_class_id':  selected_class_id,
        'all_months':         list(range(1, 13)),
        'hide_year':       hide_year,
        'hide_financial':  hide_financial,
    }

    # Build months_list cho chế độ tháng
    if selected_class_id:
        # 1) Lọc MealRecord theo lớp -> lấy distinct (year, month)
        #    Cách 1: dùng ORM annotate + values + distinct
        meals_qs = MealRecord.objects.filter(student__classroom_id=selected_class_id)
        dates = meals_qs \
            .annotate(year=ExtractYear('date'), month=ExtractMonth('date')) \
            .values('year', 'month') \
            .distinct()

        # 2) Chuyển thành list các tuple (year, month) rồi sort theo thời gian
        ym_tuples = [(d['year'], d['month']) for d in dates]
        # Sắp xếp tăng dần theo (year, month)
        ym_tuples.sort(key=lambda x: (x[0], x[1]))

        # 3) Format thành chuỗi "M/YYYY" (không có leading zero) hoặc "MM/YYYY" nếu cần
        months_list = [f"{m}/{y}" for (y, m) in ym_tuples]
        context['months_list'] = months_list
    else:
        context['months_list'] = []

    # --- Thống kê theo tháng (giữ nguyên) ---
    if mode == 'month' and selected_term and selected_month and selected_class_id:
        class_id_int = int(selected_class_id)
        students     = Student.objects.filter(classroom_id=class_id_int).order_by('name')
        m_str, y_str = selected_month.split('/')
        month_i = int(m_str)
        year_i  = int(y_str)
        max_day = calendar.monthrange(year_i, month_i)[1]

        # Tính thống kê cho cả hai loại bữa sáng và bữa trưa
        stats = {}
        for mt in ['Bữa sáng', 'Bữa trưa']:
            recs = MealRecord.objects.filter(
                student__classroom_id=class_id_int,
                meal_type=mt,
                date__year=year_i,
                date__month=month_i
            )
            record_map = {
                (r.student_id, r.date.day): {
                    'status': r.status,
                    'non_eat': r.non_eat,
                    'reason': r.absence_reason or ''
                }
                for r in recs
            }

            rows_mt = []
            totals_mt = [0] * max_day

            for stu in students:
                ym = f"{year_i}-{month_i:02d}"
                # Lấy Payment; nếu chưa có thì fee = 0
                sp = StudentPayment.objects.filter(student=stu, month=ym).first()
                if sp:
                    price     = sp.meal_price
                    fee_break = price.breakfast_price
                    fee_lunch = price.lunch_price
                else:
                    fee_break = 0
                    fee_lunch = 0

                # Với Bữa trưa: tính thêm Đã thu và Ăn Học
                if mt == 'Bữa trưa':
                    # số ngày ăn sáng và ăn trưa
                    br_count = MealRecord.objects.filter(
                        student=stu,
                        date__year=year_i, date__month=month_i,
                        meal_type='Bữa sáng'
                    ).filter(Q(status='Đủ') | Q(status='Thiếu', non_eat=2)).count()
                    lu_count = MealRecord.objects.filter(
                        student=stu,
                        date__year=year_i, date__month=month_i,
                        meal_type='Bữa trưa'
                    ).filter(Q(status='Đủ') | Q(status='Thiếu', non_eat=2)).count()
                    tuition_fee     = sp.tuition_fee if sp else 0
                    learning_cost   = int(tuition_fee) + br_count * fee_break + lu_count * fee_lunch
                    collected       = sp.amount_paid  if sp else 0

                days = []
                total_meals = 0
                total_cost  = 0

                for d in range(1, max_day + 1):
                    info = record_map.get((stu.id, d), {})
                    status = info.get('status')
                    ne     = info.get('non_eat')
                    reason = info.get('reason', '')

                    if status == 'Đủ':
                        display     = 'X'
                        total_meals += 1
                        cost_unit    = fee_break if mt == 'Bữa sáng' else fee_lunch
                        total_cost  += cost_unit
                        totals_mt[d-1] += 1
                        tooltip      = ''
                    elif status == 'Thiếu' and ne == 2:
                        display     = 'KP'
                        total_meals += 1
                        cost_unit    = fee_break if mt == 'Bữa sáng' else fee_lunch
                        total_cost  += cost_unit
                        #totals_mt[d-1] += 1
                        tooltip      = reason
                    elif status == 'Thiếu' and ne == 1:
                        display = 'P'
                        tooltip = reason
                    else:
                        display = '0'
                        tooltip = ''

                    days.append({
                        'display': display,
                        'reason': tooltip
                    })

                # Tạo 1 bản ghi duy nhất, gộp tất cả các trường
                # sau vòng for d in range(1, max_day+1): ... days.append(mark)
                if mt == 'Bữa trưa':
                    # đếm bữa sáng + bữa trưa
                    br_count = MealRecord.objects.filter(
                        student=stu, date__year=year_i, date__month=month_i,
                        meal_type='Bữa sáng'
                    ).filter(Q(status='Đủ')|Q(status='Thiếu', non_eat=2)).count()
                    lu_count = MealRecord.objects.filter(
                        student=stu, date__year=year_i, date__month=month_i,
                        meal_type='Bữa trưa'
                    ).filter(Q(status='Đủ')|Q(status='Thiếu', non_eat=2)).count()

                    tuition_fee = sp.tuition_fee if sp else 0
                    food_cost   = br_count * fee_break + lu_count * fee_lunch
                    total_due   = tuition_fee + food_cost
                    collected   = sp.amount_paid if sp else 0
                    remaining   = collected - total_due
                    if sp:
                        row_data = {
                            'student':      stu,
                            'days':         days,
                            'total_meals':  total_meals,
                            # format chuỗi có dấu phẩy
                            'food_cost':    f"{int(food_cost):,}",
                            'tuition_fee':  f"{int(tuition_fee):,}",
                            'total_due':    f"{int(total_due):,}",
                        }
                    else:
                        # chưa có payment → tất cả tiền = 0
                        row_data = {
                            'student':     stu,
                            'days':        days,
                            'total_meals': total_meals,
                            'food_cost':   0,
                            'tuition_fee': 0,
                            'total_due':   0,
                            'collected':   0,
                            'remaining':   0,
                        }
                else:
                    if sp:
                        row_data = {
                            'student':     stu,
                            'days':        days,
                            'total_meals': total_meals,
                            # format tổng tiền bữa sáng
                            'total_cost':  f"{int(total_cost):,}",
                        }
                    else:
                        # chưa có payment → total cost = 0
                        row_data = {
                            'student':     stu,
                            'days':        days,
                            'total_meals': total_meals,
                            'total_cost':  f"{0:,}",
                        }

                rows_mt.append(row_data)
                    

            # sắp theo TÊN (từ cuối của student.name)
            rows_mt = sorted(
                rows_mt,
                key=lambda r: r['student'].name.strip().split()[-1].upper()
            )
            #stats[mt] = {'rows': rows_mt, 'totals': totals_mt}
            total_meals_sum = sum(r['total_meals'] for r in rows_mt)
            if mt == 'Bữa trưa':
                # tính tổng với giá trị int (trước format)
                tot_food    = sum(int(str(r['food_cost']).replace(',', '')) for r in rows_mt)
                tot_tuition = sum(int(str(r['tuition_fee']).replace(',', '')) for r in rows_mt)
                tot_due     = sum(int(str(r['total_due']).replace(',', '')) for r in rows_mt)
                stats[mt] = {
                    'rows':                   rows_mt,
                    'totals':                 totals_mt,
                    # giữ nguyên số bữa (không cần dấu phẩy)
                    'totals_meals_sum':       total_meals_sum,
                    # format chuỗi cho tổng tiền
                    'totals_food_cost_sum':   f"{tot_food:,}",
                    'totals_tuition_fee_sum': f"{tot_tuition:,}",
                    'totals_due_sum':         f"{tot_due:,}",
                }
            else:
                tot_cost = sum(int(str(r['total_cost']).replace(',', '')) for r in rows_mt)
                stats[mt] = {
                    'rows':               rows_mt,
                    'totals':             totals_mt,
                    'totals_meals_sum':   total_meals_sum,
                    'totals_cost_sum':    f"{tot_cost:,}",
                }
        context.update({
            'stats':        stats,
            'max_day_list': list(range(1, max_day + 1)),
        })
    # --- Thống kê theo năm (mới) ---
    if mode == 'year':
        # Các param truyền lên chỉ có start_month/end_month dưới dạng "M/YYYY"
        start_str = request.GET.get('start_month')
        end_str   = request.GET.get('end_month')

        today = date.today()
        def_e_year  = today.year
        def_e_month = today.month
        sm = def_e_month - 9
        sy = def_e_year
        if sm <= 0:
            sm += 12; sy -= 1

        if start_str and end_str:
            # parse "M/YYYY" thành số tháng, năm
            m0, y0 = map(int, start_str.split('/'))
            m1, y1 = map(int, end_str.split('/'))
            start_year, start_month = y0, m0
            end_year,   end_month   = y1, m1
        else:
            # fallback: dải 9 tháng gần nhất
            start_year, start_month = sy, sm
            end_year,   end_month   = def_e_year, def_e_month

        months = []
        y, m = start_year, start_month
        while (y < end_year) or (y == end_year and m <= end_month):
            months.append((y, m))
            if m == 12: m = 1; y += 1
            else:       m += 1

        rows_year = []
        totals    = [[0,0,0] for _ in months]
        if selected_class_id:
            cid      = int(selected_class_id)
            qs       = Student.objects.filter(classroom_id=cid)
            # sort list theo token cuối của name, in uppercase để chuẩn hóa
            students = sorted(
                qs,
                key=lambda s: s.name.strip().split()[-1].upper()
            )

            for stu in students:
                data_row = []
                for idx, (yy, mm) in enumerate(months):
                    ym = f"{yy}-{mm:02d}"
                    sp = StudentPayment.objects.filter(student=stu, month=ym).first()
                    if sp:
                        paid      = sp.amount_paid
                        tuition   = sp.tuition_fee
                        remaining = sp.remaining_balance or Decimal('0')
                        price     = sp.meal_price
                        fee_b     = Decimal(price.breakfast_price)
                        fee_l     = Decimal(price.lunch_price)
                    else:
                        # nếu chưa có StudentPayment thì tất cả = 0
                        paid      = tuition = remaining = Decimal('0')
                        fee_b     = fee_l   = Decimal('0')
                    recs    = MealRecord.objects.filter(
                        student=stu, date__year=yy, date__month=mm,
                        meal_type__in=["Bữa sáng","Bữa trưa"]
                    )
                    spent_meals = sum(
                        fee_b if r.meal_type=="Bữa sáng" else fee_l
                        for r in recs
                        if (r.status=='Đủ' or (r.status=='Thiếu' and r.non_eat==2))
                    )
                    spent = tuition + spent_meals

                    totals[idx][0] += int(paid)
                    totals[idx][1] += int(spent)
                    totals[idx][2] += int(remaining)

                    data_row.append({
                        'paid':      f"{int(paid):,}",
                        'spent':     f"{int(spent):,}",
                        'remaining': f"{int(remaining):,}",
                    })
                rows_year.append({'student': stu, 'data': data_row})

        totals_fmt = [[f"{x:,}" for x in triple] for triple in totals]

        # Đưa selected_start_month/end_month về chuỗi "M/YYYY" để JS chọn đúng option
        sel_start = start_str if start_str else f"{start_month}/{start_year}"
        sel_end   = end_str   if end_str   else f"{end_month}/{end_year}"
        context.update({
            'selected_start_month': sel_start,
            'selected_end_month':   sel_end,
            'months':               [{'year': y, 'month': m} for y,m in months],
            'rows_year':            rows_year,
            'totals_year_data':     totals_fmt,
        })

    return render(request, 'meals/statistics.html', context)
def ajax_get_months(request):
    term     = request.GET.get('term')
    class_id = request.GET.get('class_id')
    if not class_id:
        return JsonResponse([], safe=False)

    # Lọc bản ghi theo lớp qua student, và nếu có term thì filter thêm
    qs = MealRecord.objects.filter(student__classroom_id=class_id)
    if term:
        qs = qs.filter(student__classroom__term=term)

    # Annotate tháng và năm từ field date
    qs = qs.annotate(
        m=ExtractMonth('date'),
        y=ExtractYear('date'),
    )

    # Lấy distinct các cặp (m,y) rồi sort theo y,m
    months = (
        qs.values_list('m', 'y')
          .distinct()
          .order_by('y', 'm')
    )

    # Build list string "M/YYYY"
    data = [f"{m}/{y}" for m, y in months]
    return JsonResponse(data, safe=False)
def ajax_load_months(request):
    class_id = request.GET.get('class_id')
    if not class_id:
        return JsonResponse([], safe=False)

    meals_qs = MealRecord.objects.filter(student__classroom_id=class_id)
    dates = meals_qs \
        .annotate(year=ExtractYear('date'), month=ExtractMonth('date')) \
        .values('year', 'month') \
        .distinct()
    ym_tuples = [(d['year'], d['month']) for d in dates]
    ym_tuples.sort(key=lambda x: (x[0], x[1]))  # sort theo (year, month)

    months_list = [f"{m}/{y}" for (y, m) in ym_tuples]
    return JsonResponse(months_list, safe=False)
import calendar
import openpyxl
from django.http import HttpResponse
from .models import Student, MealRecord, StudentPayment, ClassRoom

def export_monthly_statistics(request):
    selected_year      = request.GET.get('year')
    selected_month     = request.GET.get('month')
    selected_class_id  = request.GET.get('class_id')
    
    if not (selected_year and selected_month and selected_class_id):
        return HttpResponse("Thiếu tham số", status=400)
    
    # Chuyển kiểu và tính số ngày
    year_i  = int(selected_year)
    month_i = int(selected_month)
    max_day = calendar.monthrange(year_i, month_i)[1]
    
    # Lấy lớp và danh sách học sinh
    try:
        class_id_int = int(selected_class_id)
        classroom    = ClassRoom.objects.get(id=class_id_int)
    except:
        return HttpResponse("Lớp học không hợp lệ", status=400)
    # Lấy học sinh rồi sort theo TÊN (last word)
    students_qs = Student.objects.filter(classroom=classroom)
    students = sorted(
        students_qs,
        key=lambda s: s.name.strip().split()[-1].upper()
    )
    
    # Lấy các MealRecord trong tháng
    recs = MealRecord.objects.filter(
        student__classroom=classroom,
        meal_type=selected_meal_type,
        date__year=year_i,
        date__month=month_i
    )
    # Map (student_id, day) -> (status, non_eat)
    record_map = {
        (r.student_id, r.date.day): (r.status, r.non_eat)
        for r in recs
    }
    
    # Tạo workbook và sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"TK {selected_month}-{selected_year}"
    
    # 1. Tiêu đề merge A1→(3+max_day)2
    last_col = 3 + max_day
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=last_col)
    title = f"BẢNG CHẤM {selected_meal_type.upper()} THÁNG {selected_month}/{selected_year} - LỚP: {classroom.name}"
    cell = ws.cell(row=1, column=1, value=title)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font      = Font(size=14, bold=True)
    
    # 2. Header cột (dòng 3)
    headers = ["Học sinh"] + [str(d) for d in range(1, max_day+1)] + ["Tổng", "Thành tiền"]
    for idx, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=idx, value=h)
        c.font      = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = (12 if idx<=2 else 5)
    
    # 3. Dòng dữ liệu học sinh (bắt đầu từ row 4)
    row_num = 4
    # Dùng để tính hàng tổng sau này
    totals_per_day = [0] * max_day
    
    for stu in students:
        # Lấy daily_fee của học sinh
        ym_str = f"{selected_year}-{selected_month.zfill(2)}"
        sp = StudentPayment.objects.filter(student=stu, month=ym_str).first()
        price     = sp.meal_price if sp else get_current_price()
        fee_break = price.breakfast_price
        fee_lunch = price.lunch_price
        total_meals = 0
        total_cost  = 0
        row = [stu.name]
        
        # Chấm từng ngày
        for d in range(1, max_day+1):
            status = record_map.get((stu.id, d))
            if status:
                s, ne = status
                if s == "Đủ" or (s=="Thiếu" and ne==2):
                    mark = "X"
                    total_meals += 1
                    total_cost  += (fee_break if selected_meal_type=="Bữa sáng" else fee_lunch)
                    totals_per_day[d-1] += 1
                elif s=="Thiếu" and ne==1:
                    mark = "P"
                else:
                    mark = "0"
            else:
                mark = "0"
            row.append(mark)
        
        # Thêm cột Tổng & Thành tiền
        row += [total_meals, total_cost]
        
        # Ghi dòng
        for idx, val in enumerate(row, start=1):
            ws.cell(row=row_num, column=idx, value=val)
        row_num += 1
    
    # 4. Hàng tổng cuối cùng
    # Viết "Tổng" vào cột 1
    ws.cell(row=row_num, column=1, value="Tổng").font = Font(bold=True)
    # Viết tổng số bữa ăn mỗi ngày
    for i, cnt in enumerate(totals_per_day, start=1):
        ws.cell(row=row_num, column=1 + i, value=cnt).font = Font(bold=True)
    # Để trống 2 cột cuối (Tổng + Thành tiền)
    ws.cell(row=row_num, column=2 + max_day + 1, value="").font = Font(bold=True)
    ws.cell(row=row_num, column=3 + max_day + 1, value="").font = Font(bold=True)
    
    # 5. Trả file
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    fname = f"ThongKe_{selected_year}_{selected_month}.xlsx"
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    wb.save(resp)
    return resp

def export_yearly_statistics(request):
    # 1) Nhận param duy nhất start_month & end_month ở dạng "M/YYYY"
    start_str = request.GET.get('start_month')
    end_str   = request.GET.get('end_month')
    class_id  = request.GET.get('class_id')
    if not (start_str and end_str and class_id):
        return HttpResponse("Thiếu tham số", status=400)

    # 2) Parse "M/YYYY" → số nguyên month, year
    try:
        m0, y0 = start_str.split('/', 1)
        m1, y1 = end_str.split('/',   1)
        start_month = int(m0)
        start_year  = int(y0)
        end_month   = int(m1)
        end_year    = int(y1)
    except Exception:
        return HttpResponse("Thiếu tham số", status=400)

    # 3) Tạo list các (year, month) giữa khoảng từ start→end
    months = []
    y, m = start_year, start_month
    while (y < end_year) or (y == end_year and m <= end_month):
        months.append((y, m))
        m += 1
        if m > 12:
            m = 1; y += 1

    # 4) Lấy class & students
    classroom = get_object_or_404(ClassRoom, id=int(class_id))
    students_qs = Student.objects.filter(classroom=classroom)
    students = sorted(
        students_qs,
        key=lambda s: s.name.strip().split()[-1].upper()
    )

    # 5) Bắt đầu xuất Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    safe_title = f"{start_month:02d}-{start_year}_{end_month:02d}-{end_year}"
    ws.title = safe_title

    # Merge & Title
    last_col = 2 + len(months)*3
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=last_col)
    title = (
        f"THEO DÕI TIỀN ĂN TỪ THÁNG {start_month:02d}/{start_year} "
        f"ĐẾN THÁNG {end_month:02d}/{end_year}   |   LỚP: {classroom.name}"
    )
    c = ws.cell(row=1, column=1, value=title)
    c.alignment = Alignment('center', 'center')
    c.font      = Font(size=14, bold=True)

    # Header
    ws.cell(row=3, column=1, value="STT").font      = Font(bold=True)
    ws.cell(row=3, column=2, value="HỌ VÀ TÊN").font = Font(bold=True)
    for idx, (yy, mm) in enumerate(months):
        col = 3 + idx*3
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+2)
        h = ws.cell(row=3, column=col,   value=f"THÁNG {mm:02d}/{yy}")
        h.font = Font(bold=True); h.alignment = Alignment('center','center')
        ws.cell(row=4, column=col,   value="Đã đóng").font = Font(bold=True)
        ws.cell(row=4, column=col+1, value="Ăn học").font = Font(bold=True)
        ws.cell(row=4, column=col+2, value="Còn lại").font = Font(bold=True)

    # Tạo totals array
    totals = [[Decimal('0'), Decimal('0'), Decimal('0')] for _ in months]

    # Fill rows
    row_num = 5
    for i, stu in enumerate(students, start=1):
        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=stu.name)
        for idx, (yy, mm) in enumerate(months):
            ym = f"{yy:04d}-{mm:02d}"
            sp = StudentPayment.objects.filter(student=stu, month=ym).first()
            if sp:
                paid      = sp.amount_paid
                tuition   = sp.tuition_fee
                rem       = sp.remaining_balance or Decimal('0')
                br_count = MealRecord.objects.filter(
                    student=stu, date__year=yy, date__month=mm, meal_type="Bữa sáng"
                ).filter(Q(status='Đủ')|Q(status='Thiếu', non_eat=2)).count()
                lu_count = MealRecord.objects.filter(
                    student=stu, date__year=yy, date__month=mm, meal_type="Bữa trưa"
                ).filter(Q(status='Đủ')|Q(status='Thiếu', non_eat=2)).count()
                fee_b    = sp.meal_price.breakfast_price
                fee_l    = sp.meal_price.lunch_price
                spent    = tuition + Decimal(br_count)*Decimal(fee_b) + Decimal(lu_count)*Decimal(fee_l)
            else:
                paid = spent = rem = Decimal('0')

            totals[idx][0] += paid
            totals[idx][1] += spent
            totals[idx][2] += rem

            base = 3 + idx*3
            ws.cell(row=row_num, column=base,   value=float(paid)).number_format   = '#,##0'
            ws.cell(row=row_num, column=base+1, value=float(spent)).number_format  = '#,##0'
            ws.cell(row=row_num, column=base+2, value=float(rem)).number_format    = '#,##0'

        row_num += 1

    # Footer totals
    ws.cell(row=row_num, column=1, value="Tổng").font = Font(bold=True)
    for idx, (p, s, r) in enumerate(totals):
        base = 3 + idx*3
        ws.cell(row=row_num, column=base,   value=float(p)).font = Font(bold=True)
        ws.cell(row=row_num, column=base+1, value=float(s)).font = Font(bold=True)
        ws.cell(row=row_num, column=base+2, value=float(r)).font = Font(bold=True)

    # Gửi file về client
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp["Content-Disposition"] = (
        f'attachment; filename="TheoDoiTienAn_{classroom.name}_'
        f'{start_month:02d}{start_year}-{end_month:02d}{end_year}.xlsx"'
    )
    wb.save(resp)
    return resp

def ajax_load_meal_stats(request):
    """
    Trả về 2 object: breakfast và lunch, mỗi object có:
      - days: mảng ['X','0',...]
      - total: tổng số bữa đã ăn
      - cost: tổng tiền (số nguyên, chưa format)
    """
    student_id = request.GET.get('student_id')
    month_str  = request.GET.get('month')  # "YYYY-MM"
    data = {'breakfast': {}, 'lunch': {}}

    if student_id and month_str:
        # parse year/month
        year, month = map(int, month_str.split('-'))
        max_day = calendar.monthrange(year, month)[1]

        # lấy fee để tính thành tiền
        sp = StudentPayment.objects.filter(student_id=student_id, month=month_str).first()
        price     = sp.meal_price if sp else get_current_price()
        fee_break = Decimal(price.breakfast_price)
        fee_lunch = Decimal(price.lunch_price)

        # load tất cả MealRecord của tháng
        recs = MealRecord.objects.filter(
            student_id=student_id,
            date__year=year,
            date__month=month,
            meal_type__in=["Bữa sáng","Bữa trưa"]
        )
        # map (meal_type,day) -> (status, non_eat)
        record_map = {
            (r.meal_type, r.date.day): (r.status, r.non_eat)
            for r in recs
        }

        # build breakfast
        br_days = []
        br_count = 0
        for d in range(1, max_day+1):
            s = record_map.get(("Bữa sáng", d))
            if s and (s[0]=="Đủ" or (s[0]=="Thiếu" and s[1]==2)):
                br_days.append("X")
                br_count += 1
            else:
                br_days.append("0")
        data['breakfast'] = {
            'days': br_days,
            'total': br_count,
            'cost': int(br_count * fee_break)
        }

        # build lunch
        lu_days = []
        lu_count = 0
        for d in range(1, max_day+1):
            s = record_map.get(("Bữa trưa", d))
            if s and (s[0]=="Đủ" or (s[0]=="Thiếu" and s[1]==2)):
                lu_days.append("X")
                lu_count += 1
            else:
                lu_days.append("0")
        data['lunch'] = {
            'days': lu_days,
            'total': lu_count,
            'cost': int(lu_count * fee_lunch)
        }

    return JsonResponse(data)
def ajax_get_classes_by_term(request):
    term = request.GET.get('term')
    if not term:
        return JsonResponse([], safe=False)

    qs = ClassRoom.objects.filter(term=term).order_by('name')
    data = []
    for cls in qs:
        # Nếu bạn muốn hiển thị cả “(2025)” trong tên, đổi thành str(cls)
        data.append({
            'id': cls.id,
            'name': cls.name
        })
    return JsonResponse(data, safe=False)
def export_monthly_statistics_all(request):
    # 1) Lấy các GET-param, với fallback term→year và class→class_id
    year     = request.GET.get('year')
    term     = request.GET.get('term')
    month    = request.GET.get('month')
    class_id = request.GET.get('class_id') or request.GET.get('class')

    # 2) Nếu không có year nhưng có term, lấy phần số cuối của term
    if not year and isinstance(term, str):
        parts = term.strip().split()
        if parts:
            year = parts[-1]

    # 3) Nếu month có dấu '/', parse thành month và override year nếu cần
    month_str = str(month or '')      # luôn thành chuỗi
    if '/' in month_str:
        m, y    = month_str.split('/', 1)
        month   = m.zfill(2)          # '6' → '06'
        year    = y                   # override year nếu cần
    else:
        month = month_str  

    # 4) Kiểm tra đủ param
    if not (year and month and class_id):
        return HttpResponse("Thiếu tham số", status=400)

    # 5) Chuyển về số nguyên an toàn
    try:
        year_i      = int(year)
        month_i     = int(month)
        class_id_int = int(class_id)
    except ValueError:
        return HttpResponse("Thiếu tham số", status=400)

    # 6) Tính số ngày trong tháng
    max_day = calendar.monthrange(year_i, month_i)[1]

    # 7) Lấy đối tượng ClassRoom, danh sách học sinh
    classroom = get_object_or_404(ClassRoom, id=class_id_int)
    students_qs = Student.objects.filter(classroom=classroom)
    students = sorted(
        students_qs,
        key=lambda s: s.name.strip().split()[-1].upper()
    )

    # 8) Tạo Workbook và loop 2 sheet: "Bữa sáng" & "Bữa trưa"
    wb = openpyxl.Workbook()
    for idx_mt, mt in enumerate(["Bữa sáng", "Bữa trưa"]):
        if idx_mt == 0:
            ws = wb.active
            ws.title = mt          # ← ĐẶT TÊN cho sheet mặc định
        else:
            ws = wb.create_sheet(title=mt)

        # Title merged
        last_col = 3 + max_day + (3 if mt == "Bữa trưa" else 1)
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=2,   end_column=last_col
        )
        title = f"{mt.upper()} THÁNG {month}/{year} - LỚP {classroom.name}"
        c = ws.cell(1, 1, title)
        c.alignment = Alignment('center', 'center')
        c.font      = Font(size=14, bold=True)

        # Header row
        headers = ["Học sinh"] + [str(d) for d in range(1, max_day+1)] + ["Tổng"]
        if mt == "Bữa trưa":
            headers += ["Tiền Ăn", "Học Phí", "Thành Tiền"]
        else:
            headers += ["Thành tiền"]

        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(3, ci, h)
            cell.font      = Font(bold=True)
            cell.alignment = Alignment('center', 'center')
            ws.column_dimensions[get_column_letter(ci)].width = (12 if ci==1 else 5)

        # Data rows
        row = 4
        totals_per_day = [0]*max_day
        if mt == "Bữa sáng":
            total_cost_all = 0
        else:
            total_food_all    = 0
            total_tuition_all = 0
            total_due_all     = 0
        for stu in students:
            # Lấy giá meal & tuition
            sp = StudentPayment.objects.filter(
                student=stu, month=f"{year}-{month}"
            ).first()
            fee_br = sp.meal_price.breakfast_price if sp else get_current_price().breakfast_price
            fee_lu = sp.meal_price.lunch_price     if sp else get_current_price().lunch_price

            # Lấy record trong tháng này
            recs = MealRecord.objects.filter(
                student=stu, meal_type=mt,
                date__year=year_i, date__month=month_i
            )
            rec_map = {r.date.day:(r.status, r.non_eat) for r in recs}

            day_marks = []
            total_meals = 0
            total_cost  = 0
            for d in range(1, max_day+1):
                st = rec_map.get(d)
                if st:
                    status, ne = st
                    if status=="Đủ" or (status=="Thiếu" and ne==2):
                        mark = "X" if status=="Đủ" else "KP"
                        cnt_unit = fee_br if mt=="Bữa sáng" else fee_lu
                        total_meals += 1
                        total_cost   += cnt_unit
                        totals_per_day[d-1] += 1
                    elif status=="Thiếu" and ne==1:
                        mark = "P"
                    else:
                        mark = "0"
                else:
                    mark = "0"
                day_marks.append(mark)

            base = [stu.name] + day_marks + [total_meals]

            if mt == "Bữa trưa":
                # Tính các cột bổ sung
                tuition = sp.tuition_fee if sp else 0
                paid    = sp.amount_paid  if sp else 0
                br_ct = MealRecord.objects.filter(
                    student=stu, meal_type="Bữa sáng",
                    date__year=year_i, date__month=month_i
                ).filter(Q(status="Đủ")|Q(status="Thiếu", non_eat=2)).count()
                lu_ct = MealRecord.objects.filter(
                    student=stu, meal_type="Bữa trưa",
                    date__year=year_i, date__month=month_i
                ).filter(Q(status="Đủ")|Q(status="Thiếu", non_eat=2)).count()

                food = br_ct*fee_br + lu_ct*fee_lu
                total_due = tuition + food
                remaining = paid - total_due
                total_food_all    += food
                total_tuition_all += tuition
                total_due_all     += total_due
                row_vals = base + [food, tuition, total_due]
            else:
                # chỉ bữa sáng
                total_cost_all += total_cost
                row_vals = base + [total_cost]

            # Ghi row và format số
            for ci, v in enumerate(row_vals, start=1):
                c = ws.cell(row, ci, v)
                if isinstance(v, (int, float, Decimal)):
                    c.number_format = '#,##0'
            row += 1

        # Cuối cùng: row Tổng
        ws.cell(row, 1, "Tổng")
        # - Tổng cho từng ngày
        for idx, cnt in enumerate(totals_per_day, start=1):
            ws.cell(row, 1+idx, cnt)
        # - Tổng số ngày ăn (cột "Tổng"): tùy nếu bạn muốn in lại
        ws.cell(row, 1+max_day+1, sum(totals_per_day))
        if mt == "Bữa trưa":
            # ghi tổng Tiền Ăn, Học Phí, Thành Tiền
            c1 = ws.cell(row, 1+max_day+2, total_food_all)
            c2 = ws.cell(row, 1+max_day+3, total_tuition_all)
            c3 = ws.cell(row, 1+max_day+4, total_due_all)
    for c in (c1, c2, c3):
        c.number_format = '#,##0'
    # Trả file về client
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="TK_{classroom}_{month}.xlsx"'
    wb.save(resp)
    return resp